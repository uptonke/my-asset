#!/usr/bin/env python3
"""Fetch and normalize ETF holdings for Portfolio X-Ray.

Sources are deliberately source-specific instead of relying on one paid aggregator:
- QQQ: Invesco official JSON endpoint
- IFRA: iShares official CSV
- GRID: First Trust official holdings page
- COPX: Global X official full-holdings CSV discovered from the fund page
- VNM: VanEck official XLSX
- SRVR: CompaniesMarketCap fallback because Pacer blocks GitHub-hosted runners
- VOO / VEA: Vanguard official portfolio-holdings JSON API
- AVUV: Avantis official product page embedded ETF holdings
- USMV / PICK: iShares official CSV
- 00981A: ETFinfo public daily snapshot fallback (issuer PCF not yet directly integrated)

Each fund is fetched independently. A failed refresh does not overwrite an existing
snapshot. The summary records fresh/stale/failure state so downstream analysis can
make data-quality limits explicit.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import math
import re
import sys
import urllib.request
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


FUNDS = ("QQQ", "IFRA", "GRID", "COPX", "VNM", "SRVR", "VOO", "VEA", "AVUV", "USMV", "PICK", "00981A")
DEFAULT_OUTPUT_DIR = Path("data/holdings/latest")
HTTP_TIMEOUT = 35
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150 Safari/537.36 "
        "my-asset-holdings/1.0"
    ),
    "Accept": "*/*",
}


class HoldingsError(RuntimeError):
    pass


def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _number(value: Any) -> float | None:
    text = _clean(value)
    if not text or text in {"-", "--", "N/A", "NA"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    text = text.replace("$", "").replace(",", "").replace("%", "")
    try:
        out = float(text)
        return -out if negative else out
    except ValueError:
        return None


def _weight_decimal(value: Any) -> float | None:
    """Parse a source weight expressed in percentage points (e.g. 8.46 -> .0846)."""
    n = _number(value)
    return None if n is None else n / 100.0


def _iso_date(value: Any) -> str | None:
    text = _clean(value)
    if not text:
        return None
    # Accept full ISO timestamps used by Vanguard (for example 2026-06-30T00:00:00-04:00).
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        pass
    formats = (
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%b %d, %Y",
        "%B %d, %Y",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _extract_date(text: str) -> str | None:
    patterns = (
        r"\b(\d{1,2}/\d{1,2}/\d{4})\b",
        r"\b([A-Z][a-z]{2,8} \d{1,2}, \d{4})\b",
        r"\b(\d{4}-\d{2}-\d{2})\b",
    )
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            parsed = _iso_date(m.group(1))
            if parsed:
                return parsed
    return None


def _id_fields(preferred_type: str, preferred_value: Any, ticker: str, exchange: str = "") -> tuple[str, str]:
    ident = _clean(preferred_value)
    if ident and ident not in {"-", "--", "N/A"}:
        return preferred_type, ident
    ticker = _clean(ticker)
    exchange = _clean(exchange)
    if exchange and ticker:
        return "EXCHANGE_TICKER", f"{exchange}:{ticker}"
    return "TICKER", ticker


def _asset_class(raw: Any, ticker: str = "", name: str = "") -> str:
    text = " ".join((_clean(raw), _clean(ticker), _clean(name))).lower()
    if any(x in text for x in ("cash", "currency", "us dollar", "new taiwan dollar", "euro ", "sterling", "renminbi", "won ")):
        return "cash"
    if any(x in text for x in ("future", "option", "swap", "derivative")):
        return "derivative"
    if any(x in text for x in ("bond", "fixed income", "treasury", "note")):
        return "fixed_income"
    return "equity"


def _holding(
    *,
    ticker: Any,
    name: Any,
    weight: float | None,
    id_type: str,
    identifier: Any,
    exchange: Any = "",
    country: Any = "",
    currency: Any = "",
    asset_class: Any = "equity",
    sector: Any = "",
    shares: Any = None,
) -> dict[str, Any] | None:
    if weight is None or not math.isfinite(weight):
        return None
    ticker_s = _clean(ticker)
    name_s = _clean(name)
    exchange_s = _clean(exchange)
    id_type_s, id_s = _id_fields(id_type, identifier, ticker_s, exchange_s)
    if not (ticker_s or name_s or id_s):
        return None
    return {
        "id_type": id_type_s,
        "id": id_s,
        "ticker": ticker_s,
        "name": name_s,
        "exchange": exchange_s or None,
        "country": _clean(country) or None,
        "currency": _clean(currency) or None,
        "asset_class": _asset_class(asset_class, ticker_s, name_s),
        "sector": _clean(sector) or None,
        "weight": round(float(weight), 10),
        "shares": _number(shares),
    }


def _finalize(
    *,
    fund: str,
    as_of: str | None,
    source: str,
    source_quality: str,
    source_url: str,
    holdings: list[dict[str, Any]],
    reported_count: int | None = None,
    fallback_reason: str | None = None,
) -> dict[str, Any]:
    if not as_of:
        raise HoldingsError(f"{fund}: source date missing")
    holdings = [h for h in holdings if h]
    if not holdings:
        raise HoldingsError(f"{fund}: no holdings parsed")

    total_weight = sum(h["weight"] for h in holdings)
    equity_weight = sum(h["weight"] for h in holdings if h["asset_class"] == "equity")
    as_of_date = date.fromisoformat(as_of)
    stale_days = max(0, (datetime.now(timezone.utc).date() - as_of_date).days)

    # Sanity bounds catch HTML/CSV parser drift before a bad file replaces cache.
    if len(holdings) < 10:
        raise HoldingsError(f"{fund}: suspiciously low row count {len(holdings)}")
    if not 0.80 <= total_weight <= 1.20:
        raise HoldingsError(f"{fund}: parsed weight sum {total_weight:.4f} outside sanity range")

    return {
        "schema_version": 1,
        "fund": fund,
        "as_of": as_of,
        "fetched_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "source": source,
        "source_quality": source_quality,
        "source_url": source_url,
        "lookthrough_mode": "EQUITY_LOOKTHROUGH",
        "coverage_pct": round(total_weight * 100.0, 4),
        "equity_coverage_pct": round(equity_weight * 100.0, 4),
        "stale_days": stale_days,
        "holdings_count": len(holdings),
        "reported_holdings_count": reported_count,
        "fallback_reason": fallback_reason,
        "holdings": holdings,
    }


def fetch_qqq(session: requests.Session) -> dict[str, Any]:
    url = (
        "https://dng-api.invesco.com/cache/v1/accounts/en_US/shareclasses/QQQ/"
        "holdings/fund?idType=ticker&interval=monthly&productType=ETF"
    )
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 my-asset-holdings/1.0",
            "Accept": "application/json,text/plain,*/*",
            "Referer": "https://www.invesco.com/qqq-etf/en/about.html",
        },
    )
    with urllib.request.urlopen(req, timeout=HTTP_TIMEOUT) as resp:
        data = json.loads(resp.read().decode("utf-8", errors="replace"))
    raw_rows = data.get("holdings") or []
    holdings: list[dict[str, Any]] = []
    for row in raw_rows:
        h = _holding(
            ticker=row.get("ticker"),
            name=row.get("issuerName"),
            weight=_weight_decimal(row.get("percentageOfTotalNetAssets")),
            id_type="CUSIP",
            identifier=row.get("cusip"),
            currency=row.get("currency") or row.get("localCurrencyName"),
            asset_class=row.get("securityTypeName"),
            shares=row.get("units"),
        )
        if h:
            holdings.append(h)
    as_of = _iso_date(data.get("effectiveBusinessDate")) or _iso_date(data.get("effectiveDate"))
    return _finalize(
        fund="QQQ",
        as_of=as_of,
        source="Invesco official DNG holdings API",
        source_quality="OFFICIAL_DAILY",
        source_url=url,
        holdings=holdings,
        reported_count=int(data.get("totalNumberOfHoldings")) if str(data.get("totalNumberOfHoldings", "")).isdigit() else None,
    )


def _fetch_ishares_csv(
    session: requests.Session,
    *,
    fund: str,
    product_id: str,
    slug: str,
) -> dict[str, Any]:
    url = f"https://www.ishares.com/us/products/{product_id}/{slug}/latest-holdings.csv"
    r = session.get(url, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    text = r.content.decode("utf-8-sig", errors="replace")
    lines = text.splitlines()
    try:
        header_idx = next(i for i, line in enumerate(lines) if line.startswith("Ticker,"))
    except StopIteration as exc:
        raise HoldingsError(f"{fund}: iShares CSV header not found") from exc
    as_of = _extract_date(" ".join(lines[:header_idx]))
    rows = csv.DictReader(io.StringIO("\n".join(lines[header_idx:])))
    holdings: list[dict[str, Any]] = []
    for row in rows:
        ticker = row.get("Ticker", "")
        name = row.get("Name", "")
        weight = _weight_decimal(row.get("Weight (%)"))
        if weight is None:
            continue
        h = _holding(
            ticker=ticker,
            name=name,
            weight=weight,
            id_type="EXCHANGE_TICKER",
            identifier="",
            exchange=row.get("Exchange"),
            country=row.get("Location"),
            currency=row.get("Currency") or row.get("Market Currency"),
            asset_class=row.get("Asset Class"),
            sector=row.get("Sector"),
            shares=row.get("Quantity"),
        )
        if h:
            holdings.append(h)
    return _finalize(
        fund=fund,
        as_of=as_of,
        source="iShares official holdings CSV",
        source_quality="OFFICIAL_DAILY",
        source_url=url,
        holdings=holdings,
    )


def fetch_ifra(session: requests.Session) -> dict[str, Any]:
    return _fetch_ishares_csv(
        session,
        fund="IFRA",
        product_id="294315",
        slug="ishares-u-s-infrastructure-etf",
    )


def fetch_usmv(session: requests.Session) -> dict[str, Any]:
    return _fetch_ishares_csv(
        session,
        fund="USMV",
        product_id="239695",
        slug="ishares-msci-usa-minimum-volatility-etf",
    )


def fetch_pick(session: requests.Session) -> dict[str, Any]:
    return _fetch_ishares_csv(
        session,
        fund="PICK",
        product_id="239655",
        slug="ishares-msci-global-metals-mining-producers-etf",
    )


def fetch_grid(session: requests.Session) -> dict[str, Any]:
    url = "https://www.ftportfolios.com/Retail/Etf/EtfHoldings.aspx?Ticker=GRID"
    r = session.get(url, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    page_text = soup.get_text(" ", strip=True)
    m = re.search(r"Holdings of the Fund as of\s+([0-9/]+)", page_text, re.I)
    as_of = _iso_date(m.group(1)) if m else None
    count_m = re.search(r"Total Number of Holdings \(excluding cash\):\s*(\d+)", page_text, re.I)
    reported_count = int(count_m.group(1)) if count_m else None

    holdings: list[dict[str, Any]] = []
    for tr in soup.find_all("tr"):
        # recursive=False prevents page-layout tables from swallowing the holdings table.
        cells = [_clean(c.get_text(" ", strip=True)) for c in tr.find_all(["td", "th"], recursive=False)]
        if len(cells) != 7:
            continue
        name, ticker, cusip, classification, shares, _market_value, weight_text = cells
        weight = _weight_decimal(weight_text)
        if weight is None or name.lower() == "security name":
            continue
        # Data rows have a security identifier/ticker plus a percent weighting.
        if not re.fullmatch(r"-?\d+(?:\.\d+)?%", weight_text.replace(" ", "")):
            continue
        h = _holding(
            ticker=ticker,
            name=name,
            weight=weight,
            id_type="CUSIP",
            identifier=cusip,
            asset_class="cash" if ticker.startswith("$") or classification.lower() == "other" else "equity",
            sector=classification,
            shares=shares,
        )
        if h:
            holdings.append(h)
    return _finalize(
        fund="GRID",
        as_of=as_of,
        source="First Trust official holdings page",
        source_quality="OFFICIAL_DAILY",
        source_url=url,
        holdings=holdings,
        reported_count=reported_count,
    )


def fetch_copx(session: requests.Session) -> dict[str, Any]:
    page_url = "https://www.globalxetfs.com/funds/COPX"
    p = session.get(page_url, timeout=HTTP_TIMEOUT)
    p.raise_for_status()
    soup = BeautifulSoup(p.text, "html.parser")
    csv_links: list[str] = []
    for a in soup.find_all("a", href=True):
        href = str(a["href"])
        if "full-holdings" in href.lower() and href.lower().endswith(".csv"):
            csv_links.append(urljoin(page_url, href))
    if not csv_links:
        csv_links = re.findall(
            r'https?://[^"\']*copx[^"\']*full-holdings[^"\']*\.csv',
            p.text,
            flags=re.I,
        )
    if not csv_links:
        raise HoldingsError("COPX: official full-holdings CSV link not found")
    csv_url = csv_links[0]
    r = session.get(csv_url, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    text = r.content.decode("utf-8-sig", errors="replace")
    lines = text.splitlines()
    try:
        header_idx = next(i for i, line in enumerate(lines) if line.startswith("% of Net Assets,Ticker,Name,"))
    except StopIteration as exc:
        raise HoldingsError("COPX: CSV header not found") from exc
    as_of = _extract_date(" ".join(lines[:header_idx])) or _extract_date(csv_url)
    rows = csv.DictReader(io.StringIO("\n".join(lines[header_idx:])))
    holdings: list[dict[str, Any]] = []
    for row in rows:
        h = _holding(
            ticker=row.get("Ticker"),
            name=row.get("Name"),
            weight=_weight_decimal(row.get("% of Net Assets")),
            id_type="SEDOL",
            identifier=row.get("SEDOL"),
            asset_class="equity",
            shares=row.get("Shares Held"),
        )
        if h:
            holdings.append(h)
    return _finalize(
        fund="COPX",
        as_of=as_of,
        source="Global X official full holdings CSV",
        source_quality="OFFICIAL_DAILY",
        source_url=csv_url,
        holdings=holdings,
    )


def fetch_vnm(session: requests.Session) -> dict[str, Any]:
    url = "https://www.vaneck.com/us/en/investments/vietnam-etf-vnm/downloads/holdings/"
    r = session.get(url, timeout=HTTP_TIMEOUT, allow_redirects=True)
    r.raise_for_status()
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    as_of = _extract_date(" ".join(_clean(x) for row in rows[:8] for x in row if x is not None))
    if not as_of:
        as_of = _extract_date(wb.sheetnames[0].replace("_", "/"))

    header_idx = None
    headers: list[str] = []
    for idx, row in enumerate(rows):
        normalized = [_clean(x) for x in row]
        if "Ticker" in normalized and any("Net Assets" in x for x in normalized):
            header_idx = idx
            headers = normalized
            break
    if header_idx is None:
        raise HoldingsError("VNM: holdings header row not found")
    index = {name: i for i, name in enumerate(headers) if name}

    def cell(row: list[Any], key: str) -> Any:
        i = index.get(key)
        return row[i] if i is not None and i < len(row) else None

    weight_header = next((h for h in headers if "Net Assets" in h), "% of Net Assets")
    holdings: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        ticker = cell(row, "Ticker")
        name = cell(row, "Holding Name")
        weight = _weight_decimal(cell(row, weight_header))
        if weight is None:
            continue
        h = _holding(
            ticker=ticker,
            name=name,
            weight=weight,
            id_type="FIGI",
            identifier=cell(row, "Identifier (FIGI)"),
            asset_class=cell(row, "Asset Class"),
            shares=cell(row, "Shares"),
        )
        if h:
            holdings.append(h)
    return _finalize(
        fund="VNM",
        as_of=as_of,
        source="VanEck official holdings XLSX",
        source_quality="OFFICIAL_DAILY",
        source_url=url,
        holdings=holdings,
    )



def _fetch_vanguard(session: requests.Session, fund: str) -> dict[str, Any]:
    base_url = f"https://investor.vanguard.com/investment-products/etfs/profile/api/{fund}/portfolio-holding/stock"
    start = 1
    page_size = 500
    raw_rows: list[dict[str, Any]] = []
    as_of: str | None = None
    reported_count: int | None = None

    while True:
        r = session.get(base_url, params={"start": start, "count": page_size}, timeout=HTTP_TIMEOUT)
        r.raise_for_status()
        data = r.json()
        if as_of is None:
            as_of = _iso_date(data.get("asOfDate"))
        try:
            reported_count = int(data.get("size"))
        except (TypeError, ValueError):
            reported_count = None
        rows = ((data.get("fund") or {}).get("entity") or [])
        if not isinstance(rows, list) or not rows:
            break
        raw_rows.extend(row for row in rows if isinstance(row, dict))
        if (reported_count is not None and len(raw_rows) >= reported_count) or len(rows) < page_size:
            break
        start += len(rows)
        if start > 10000:
            raise HoldingsError(f"{fund}: Vanguard pagination exceeded safety limit")

    holdings: list[dict[str, Any]] = []
    for row in raw_rows:
        ident = row.get("isin") or row.get("cusip") or row.get("sedol")
        ident_type = "ISIN" if row.get("isin") else ("CUSIP" if row.get("cusip") else "SEDOL")
        h = _holding(
            ticker=row.get("ticker"),
            name=row.get("longName") or row.get("shortName"),
            weight=_weight_decimal(row.get("percentWeight")),
            id_type=ident_type,
            identifier=ident,
            asset_class="equity",
            shares=row.get("sharesHeld"),
        )
        if h:
            holdings.append(h)

    return _finalize(
        fund=fund,
        as_of=as_of,
        source="Vanguard official portfolio holdings API",
        source_quality="OFFICIAL_MONTHLY",
        source_url=base_url,
        holdings=holdings,
        reported_count=reported_count,
    )


def fetch_voo(session: requests.Session) -> dict[str, Any]:
    return _fetch_vanguard(session, "VOO")


def fetch_vea(session: requests.Session) -> dict[str, Any]:
    return _fetch_vanguard(session, "VEA")


def _extract_balanced_js_array(text: str, marker: str) -> str:
    marker_idx = text.find(marker)
    if marker_idx < 0:
        raise HoldingsError(f"embedded array marker not found: {marker}")
    start = text.find("[", marker_idx + len(marker) - 1)
    if start < 0:
        raise HoldingsError(f"embedded array opening bracket not found: {marker}")
    depth = 0
    in_string = False
    escaped = False
    for idx in range(start, len(text)):
        ch = text[idx]
        if in_string:
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                return text[start : idx + 1]
    raise HoldingsError(f"unterminated embedded array: {marker}")


def _split_js_object_array(array_text: str) -> list[str]:
    objects: list[str] = []
    depth = 0
    start: int | None = None
    in_string = False
    escaped = False
    for idx, ch in enumerate(array_text):
        if in_string:
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch == "{":
            if depth == 0:
                start = idx
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and start is not None:
                objects.append(array_text[start : idx + 1])
                start = None
    return objects


def _js_string_field(obj: str, key: str) -> str:
    m = re.search(rf'(?:^|,)\s*{re.escape(key)}:"((?:\\.|[^"\\])*)"', obj)
    if not m:
        return ""
    try:
        return json.loads('"' + m.group(1) + '"')
    except Exception:
        return m.group(1)


def fetch_avuv(session: requests.Session) -> dict[str, Any]:
    url = "https://www.avantisinvestors.com/avantis-investments/avantis-us-small-cap-value-etf/"
    r = session.get(url, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    html = r.text
    as_of_match = re.search(r'etfHoldingsAsOfDate:"([^"]+)"', html)
    as_of = _iso_date(as_of_match.group(1)) if as_of_match else None
    count_match = re.search(r'numberOfHoldings:"([\d,]+)"', html)
    reported_count = int(count_match.group(1).replace(",", "")) if count_match else None

    array_text = _extract_balanced_js_array(html, "etfHoldings:[")
    holdings: list[dict[str, Any]] = []
    for obj in _split_js_object_array(array_text):
        ticker = _js_string_field(obj, "ticker")
        name = _js_string_field(obj, "name")
        security_type = _js_string_field(obj, "securityType")
        isin = _js_string_field(obj, "isin")
        cusip = _js_string_field(obj, "cusip")
        weight = _weight_decimal(_js_string_field(obj, "weight"))
        h = _holding(
            ticker=ticker,
            name=name,
            weight=weight,
            id_type="ISIN" if isin else "CUSIP",
            identifier=isin or cusip,
            country=_js_string_field(obj, "country"),
            asset_class=security_type,
            sector=_js_string_field(obj, "sector"),
            shares=_js_string_field(obj, "shareQuantity"),
        )
        if h:
            holdings.append(h)

    return _finalize(
        fund="AVUV",
        as_of=as_of,
        source="Avantis official product page embedded ETF holdings",
        source_quality="OFFICIAL_DAILY",
        source_url=url,
        holdings=holdings,
        reported_count=reported_count,
    )


def _nuxt_ref(payload: list[Any], ref: Any) -> Any:
    if isinstance(ref, int) and not isinstance(ref, bool) and 0 <= ref < len(payload):
        return payload[ref]
    return ref


def fetch_00981a(session: requests.Session) -> dict[str, Any]:
    # The issuer's machine-readable daily PCF is not yet integrated. ETFinfo exposes a
    # public server-rendered snapshot sourced from Taiwan market disclosures; keep it
    # explicitly labelled as a third-party fallback.
    url = "https://www.etfinfo.tw/etf/00981A/holdings"
    r = session.get(url, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    payload: list[Any] | None = None
    for script in soup.find_all("script"):
        text = script.string or script.get_text("", strip=False) or ""
        if "etf-detail-base-00981A" not in text or "ShallowReactive" not in text:
            continue
        try:
            candidate = json.loads(text)
        except Exception:
            continue
        if isinstance(candidate, list):
            payload = candidate
            break
    if not payload:
        raise HoldingsError("00981A: Nuxt holdings payload not found")

    root_tag = payload[0] if payload else None
    if not (isinstance(root_tag, list) and len(root_tag) >= 2):
        raise HoldingsError("00981A: unexpected Nuxt root")
    root = _nuxt_ref(payload, root_tag[1])
    data_tag = _nuxt_ref(payload, root.get("data") if isinstance(root, dict) else None)
    if isinstance(data_tag, list) and len(data_tag) >= 2 and data_tag[0] == "ShallowReactive":
        data = _nuxt_ref(payload, data_tag[1])
    else:
        data = data_tag
    if not isinstance(data, dict):
        raise HoldingsError("00981A: Nuxt data object missing")
    detail = _nuxt_ref(payload, data.get("etf-detail-base-00981A"))
    if not isinstance(detail, dict):
        raise HoldingsError("00981A: ETF detail object missing")

    latest_market = _nuxt_ref(payload, detail.get("latestMarket"))
    as_of = None
    if isinstance(latest_market, dict):
        as_of = _iso_date(_nuxt_ref(payload, latest_market.get("date")))
    holdings_snapshot = _nuxt_ref(payload, detail.get("holdings"))
    if not isinstance(holdings_snapshot, dict):
        raise HoldingsError("00981A: holdings snapshot object missing")
    # Nuxt stores the snapshot object separately from its inner holdings array.
    # detail["holdings"] -> snapshot dict -> snapshot["holdings"] -> list of row refs.
    raw_holdings = _nuxt_ref(payload, holdings_snapshot.get("holdings"))
    if not isinstance(raw_holdings, list):
        raise HoldingsError("00981A: holdings list missing")
    if not as_of:
        as_of = _iso_date(_nuxt_ref(payload, holdings_snapshot.get("snapshotDate")))

    holdings: list[dict[str, Any]] = []
    for href in raw_holdings:
        obj = _nuxt_ref(payload, href)
        if not isinstance(obj, dict):
            continue
        ticker = _clean(_nuxt_ref(payload, obj.get("code")))
        name = _clean(_nuxt_ref(payload, obj.get("name")))
        weight_raw = _nuxt_ref(payload, obj.get("weight"))
        shares_raw = _nuxt_ref(payload, obj.get("shares"))
        industry = _clean(_nuxt_ref(payload, obj.get("industry")))
        is_derivative = ticker == "TX" or "期貨" in name or "FUTURE" in name.upper()
        h = _holding(
            ticker=ticker,
            name=name,
            weight=_weight_decimal(weight_raw),
            id_type="TICKER",
            identifier=ticker,
            country="Taiwan",
            currency="TWD",
            asset_class="derivative" if is_derivative else "equity",
            sector=industry,
            shares=shares_raw,
        )
        if h:
            holdings.append(h)

    return _finalize(
        fund="00981A",
        as_of=as_of,
        source="ETFinfo public holdings snapshot fallback",
        source_quality="THIRD_PARTY_FALLBACK",
        source_url=url,
        holdings=holdings,
        reported_count=len(raw_holdings),
        fallback_reason="Issuer daily PCF is not directly integrated; ETFinfo public snapshot is used as a transparent fallback.",
    )

def fetch_srvr(session: requests.Session) -> dict[str, Any]:
    # Pacer's official site currently returns Cloudflare 403 from GitHub-hosted runners.
    # Keep this fallback explicitly lower-quality rather than pretending it is issuer data.
    url = "https://companiesmarketcap.com/pacer-benchmark-data-infrastructure-real-estate-sctr-etf/holdings/"
    r = session.get(url, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    page_text = soup.get_text(" ", strip=True)
    m = re.search(r"Etf holdings as of ([A-Za-z]+ \d{1,2}, \d{4})", page_text, re.I)
    as_of = _iso_date(m.group(1)) if m else None
    count_m = re.search(r"Number of holdings:\s*(\d+)", page_text, re.I)
    reported_count = int(count_m.group(1)) if count_m else None

    holdings: list[dict[str, Any]] = []
    for tr in soup.find_all("tr"):
        cells = [_clean(c.get_text(" ", strip=True)) for c in tr.find_all(["td", "th"], recursive=False)]
        if len(cells) < 4:
            continue
        weight_text, name, ticker, shares = cells[:4]
        if not re.fullmatch(r"-?\d+(?:\.\d+)?%", weight_text.replace(" ", "")):
            continue
        h = _holding(
            ticker=ticker,
            name=name,
            weight=_weight_decimal(weight_text),
            id_type="TICKER",
            identifier=ticker,
            asset_class="equity",
            shares=shares,
        )
        if h:
            holdings.append(h)
    return _finalize(
        fund="SRVR",
        as_of=as_of,
        source="CompaniesMarketCap holdings fallback",
        source_quality="THIRD_PARTY_FALLBACK",
        source_url=url,
        holdings=holdings,
        reported_count=reported_count,
        fallback_reason="Pacer official holdings page is blocked by Cloudflare on GitHub-hosted runners.",
    )


FETCHERS = {
    "QQQ": fetch_qqq,
    "IFRA": fetch_ifra,
    "GRID": fetch_grid,
    "COPX": fetch_copx,
    "VNM": fetch_vnm,
    "SRVR": fetch_srvr,
    "VOO": fetch_voo,
    "VEA": fetch_vea,
    "AVUV": fetch_avuv,
    "USMV": fetch_usmv,
    "PICK": fetch_pick,
    "00981A": fetch_00981a,
}


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tmp.replace(path)


def _snapshot_materially_equal(old: Any, new: Any) -> bool:
    if not isinstance(old, dict) or not isinstance(new, dict):
        return False
    ignored = {"fetched_at", "stale_days"}
    old_cmp = {k: v for k, v in old.items() if k not in ignored}
    new_cmp = {k: v for k, v in new.items() if k not in ignored}
    return old_cmp == new_cmp


def run(output_dir: Path, strict: bool = False, funds: tuple[str, ...] = FUNDS) -> int:
    session = _session()
    summary: dict[str, Any] = {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "funds": {},
    }
    hard_failures = 0

    for fund in funds:
        fetcher = FETCHERS[fund]
        path = output_dir / f"{fund}.json"
        try:
            payload = fetcher(session)
            cached_before = None
            if path.exists():
                try:
                    cached_before = json.loads(path.read_text(encoding="utf-8"))
                except Exception:
                    cached_before = None
            if not _snapshot_materially_equal(cached_before, payload):
                _write_json(path, payload)
            summary["funds"][fund] = {
                "status": "FRESH",
                "as_of": payload["as_of"],
                "stale_days": payload["stale_days"],
                "source_quality": payload["source_quality"],
                "holdings_count": payload["holdings_count"],
                "coverage_pct": payload["coverage_pct"],
                "equity_coverage_pct": payload["equity_coverage_pct"],
            }
            print(
                f"✅ {fund}: {payload['holdings_count']} rows, "
                f"coverage={payload['coverage_pct']:.2f}%, as_of={payload['as_of']}, "
                f"source={payload['source_quality']}",
                flush=True,
            )
        except Exception as exc:
            cached = None
            if path.exists():
                try:
                    cached = json.loads(path.read_text(encoding="utf-8"))
                except Exception:
                    cached = None
            if cached:
                summary["funds"][fund] = {
                    "status": "STALE_CACHE",
                    "as_of": cached.get("as_of"),
                    "stale_days": cached.get("stale_days"),
                    "source_quality": cached.get("source_quality"),
                    "holdings_count": cached.get("holdings_count"),
                    "coverage_pct": cached.get("coverage_pct"),
                    "equity_coverage_pct": cached.get("equity_coverage_pct"),
                    "refresh_error": f"{type(exc).__name__}: {exc}",
                }
                print(f"⚠️ {fund}: refresh failed; retained cached snapshot: {exc}", flush=True)
                if strict:
                    hard_failures += 1
            else:
                hard_failures += 1
                summary["funds"][fund] = {
                    "status": "FAILED_NO_CACHE",
                    "refresh_error": f"{type(exc).__name__}: {exc}",
                }
                print(f"❌ {fund}: no usable snapshot: {exc}", flush=True)

    statuses = [v.get("status") for v in summary["funds"].values()]
    summary["fresh_count"] = statuses.count("FRESH")
    summary["stale_cache_count"] = statuses.count("STALE_CACHE")
    summary["failed_count"] = statuses.count("FAILED_NO_CACHE")
    _write_json(output_dir / "summary.json", summary)

    if hard_failures:
        print(f"Holdings ingestion completed with {hard_failures} hard failure(s).", file=sys.stderr)
        return 1
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--strict", action="store_true", help="Fail if any requested fresh fetch fails.")
    parser.add_argument("--fund", action="append", choices=FUNDS, help="Fetch only selected fund(s).")
    args = parser.parse_args()
    funds = tuple(args.fund) if args.fund else FUNDS
    return run(args.output_dir, strict=args.strict, funds=funds)


if __name__ == "__main__":
    raise SystemExit(main())
